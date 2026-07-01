import openpyxl, json, datetime, os, re

BASE = '..'
OUT = '../app/src/data'
os.makedirs(OUT, exist_ok=True)

def s(v):
    if v is None:
        return None
    if isinstance(v, datetime.datetime):
        return v.strftime('%Y-%m-%d')
    if isinstance(v, datetime.time):
        return v.strftime('%H:%M')
    if isinstance(v, datetime.date):
        return v.strftime('%Y-%m-%d')
    if isinstance(v, str):
        return v.strip()
    return v

# ---------- Incident register ----------
wb = openpyxl.load_workbook(f'{BASE}/Журнал НС 16.06.26.xlsx', read_only=True, data_only=True)
ws = wb['НС 2020-2025']
rows = list(ws.iter_rows(values_only=True))
header = rows[1]

cols = {
 'production': 0, 'status': 1, 'num': 2, 'group': 3, 'date': 4, 'weekday': 5,
 'time': 6, 'period': 7, 'classification': 8, 'dzo': 9, 'businessLine': 10,
 'region': 11, 'place': 12, 'type': 13, 'description': 14, 'firstMeasures': 15,
 'liquidationDate': 16, 'liquidationTime': 17, 'productionImpact': 18, 'notifiedBodies': 19,
 'prelimCauses': 20, 'consequences': 21, 'damage': 22, 'mainCauses': 23, 'rootCauses': 24,
 'workType': 25, 'zone': 26, 'victim': 27, 'gender': 28, 'birthYear': 29, 'age': 30,
 'ageCategory': 31, 'position': 32, 'personnelCategory': 33, 'severity': 34, 'diagnosis': 35,
 'bodyPart': 36, 'employerFault': 37, 'workerFault': 38, 'tenure': 39, 'tenureCategory': 40,
}

def norm_severity(v):
    if not v: return None
    t = str(v).lower()
    if 'летал' in t: return 'fatal'
    if 'тяж' in t: return 'severe'
    if 'легк' in t: return 'light'
    return 'other'

def norm_dzo(v):
    if not v: return None
    t = re.sub(r'[«»"\']', '', str(v)).strip().lower()
    t = re.sub(r'\s+', ' ', t)
    mapping = [
        ('озенмунайгаз', 'АО «Озенмунайгаз»'),
        ('каражанбасмунай', 'АО «Каражанбасмунай»'),
        ('озенмунайсервис', 'ТОО «ОзенМунайСервис»'),
        ('oil services', 'ТОО «Oil Services Company»'),
        ('oil construction', 'ТОО «Oil Construction Company»'),
        ('казтрансойл', 'АО «КазТрансОйл»'),
        ('эмбамунайгаз', 'АО «Эмбамунайгаз»'),
        ('мангистаумунайгаз', 'АО «Мангистаумунайгаз»'),
        ('kmg international', 'KMG International'),
        ('kmg-security', 'ТОО «KMG-Security»'),
        ('kmg security', 'ТОО «KMG-Security»'),
        ('кен курылыс', 'ТОО «Кен Курылыс Сервис»'),
        ('казахойл актобе', 'ТОО «КазахОйл Актобе»'),
        ('казойл актобе', 'ТОО «КазахОйл Актобе»'),
        ('петроказахстан ойл', 'ТОО «ПетроКазахстан Ойл Продактс»'),
        ('ep-catering', 'ТОО «KMG EP-Catering»'),
        ('ep- catering', 'ТОО «KMG EP-Catering»'),
        ('kpi inc', 'ТОО «KPI Inc.»'),
        ('атырауский нефтепер', 'ТОО «Атырауский НПЗ»'),
        ('павлодарский нефтехим', 'ТОО «ПНХЗ»'),
        ('тулпармунайсервис', 'ТОО «ТулпарМунайСервис»'),
        ('казгермунай', 'ТОО СП «КазГерМунай»'),
        ('казахский газопер', 'ТОО «КазГПЗ»'),
        ('казмунайгаз', 'АО НК «КазМунайГаз»'),
        ('кмг инжиниринг', 'ТОО «КМГ Инжиниринг»'),
        ('удтв', 'ТОО «УДТВ»'),
        ('transport corporation', 'ТОО «Oil Transport Corporation»'),
        ('argymaktransservice', 'ТОО «ArgymakTransService»'),
    ]
    for key, name in mapping:
        if key in t:
            return name
    return str(v).strip()

def norm_type(v):
    if not v: return None
    t = str(v).strip().rstrip(';').strip().lower()
    if 'высот' in t and 'паден' in t: return 'fall_height'
    if 'паден' in t and 'пострадав' in t: return 'fall'
    if 'движущ' in t or 'вращающ' in t: return 'moving_parts'
    if 'обруш' in t or 'обвал' in t: return 'collapse'
    if 'дтп' in t or 'дорожн' in t or 'транспортн' in t: return 'vehicle'
    if 'электр' in t: return 'electric'
    if 'температур' in t or 'пожар' in t: return 'temperature'
    if 'здоров' in t or 'сердеч' in t: return 'health'
    if 'вредн' in t or 'опасн' in t: return 'hazardous'
    return 'other'

incidents = []
for r in rows[2:]:
    if r[cols['num']] is None and r[cols['date']] is None and r[cols['victim']] is None:
        continue
    rec = {}
    for k, idx in cols.items():
        rec[k] = s(r[idx])
    d = rec.get('date')
    if isinstance(d, str) and len(d) >= 4 and d[:4].isdigit():
        rec['year'] = int(d[:4])
        rec['month'] = int(d[5:7])
    else:
        rec['year'] = None
        rec['month'] = None
    rec['severityNorm'] = norm_severity(rec.get('severity'))
    rec['dzoNorm'] = norm_dzo(rec.get('dzo'))
    rec['typeNorm'] = norm_type(rec.get('type'))
    rec['id'] = len(incidents) + 1
    incidents.append(rec)

with open(f'{OUT}/incidents.json', 'w', encoding='utf-8') as f:
    json.dump(incidents, f, ensure_ascii=False, indent=0)
print('incidents:', len(incidents))

# ---------- Classifier (bilingual incident types) ----------
ws = wb['Классификатор НС']
crows = list(ws.iter_rows(values_only=True))
types = []
for r in crows[1:22]:
    ru = s(r[0]); kz = s(r[1])
    if ru:
        counts = {y: (r[i] or 0) for y, i in [(2023,2),(2022,3),(2021,4),(2020,5)]}
        types.append({'ru': ru, 'kz': kz, 'counts': counts})
with open(f'{OUT}/classifier.json', 'w', encoding='utf-8') as f:
    json.dump(types, f, ensure_ascii=False, indent=0)
print('classifier types:', len(types))

# ---------- Korgau by DZO ----------
wb2 = openpyxl.load_workbook(f'{BASE}/Коргау_EBD_Расширенная_База_2023_2026.xlsx', read_only=True, data_only=True)
ws = wb2['По ДЗО']
krows = list(ws.iter_rows(values_only=True))
korgau = []
for r in krows[2:]:
    if r[0] is None or str(r[0]).startswith('Коргау'):
        continue
    korgau.append({
        'org': s(r[0]), 'total': r[1], 'nm': r[2], 'nm1000': r[3],
        'unresolved': r[4], 'gp': r[5],
        'y2023': r[6], 'y2024': r[7], 'y2025': r[8], 'y2026': r[9],
    })
with open(f'{OUT}/korgau.json', 'w', encoding='utf-8') as f:
    json.dump(korgau, f, ensure_ascii=False, indent=0)
print('korgau orgs:', len(korgau))

# korgau summary
ws = wb2['Сводная']
summ = {}
for r in ws.iter_rows(values_only=True):
    if r[0] and r[1] is not None:
        summ[str(r[0])] = r[1]
with open(f'{OUT}/korgau_summary.json', 'w', encoding='utf-8') as f:
    json.dump(summ, f, ensure_ascii=False, indent=0)

# ---------- Predictive forecast ----------
wb3 = openpyxl.load_workbook(f'{BASE}/Коргау_ИНР_Предиктивная_Аналитика_2026 (3).xlsx', read_only=True, data_only=True)
ws = wb3['Прогноз Q2-Q3 все ДЗО']
frows = list(ws.iter_rows(values_only=True))
forecast = []
for r in frows[2:]:
    dzo = s(r[0])
    if not dzo or dzo.startswith('▌') or dzo.startswith('КРАСНАЯ'):
        continue
    if r[1] is None:
        continue
    forecast.append({
        'dzo': dzo, 'quarter': s(r[1]), 'risk': s(r[2]), 'work': s(r[3]),
        'probability': s(r[4]), 'precursors': s(r[5]), 'history': s(r[6]), 'action': s(r[7]),
    })
with open(f'{OUT}/forecast.json', 'w', encoding='utf-8') as f:
    json.dump(forecast, f, ensure_ascii=False, indent=0)
print('forecast rows:', len(forecast))

# ---------- Electronic permit journal (PTW / END) ----------
def ptw_work_category(text):
    if not text or text == '—':
        return 'other'
    t = str(text).lower()
    if 'огнев' in t or 'отты' in t or 'сварк' in t or 'дәнекер' in t:
        return 'hot_work'
    if 'замкнут' in t or 'ограничен' in t or 'тұйық' in t:
        return 'confined'
    if 'высот' in t or 'биікт' in t:
        return 'height'
    if 'землян' in t or 'жер ' in t:
        return 'excavation'
    if 'обустрой' in t or 'жайластыру' in t:
        return 'well_completion'
    if 'парафин' in t or 'продув' in t:
        return 'paraffin'
    if 'качал' in t or 'станок' in t or 'тербел' in t:
        return 'pumping_unit'
    if 'электр' in t or 'loto' in t:
        return 'electrical'
    return 'other'

def ptw_risk(cat):
    if cat in ('hot_work', 'confined', 'height'):
        return 'high'
    if cat in ('excavation', 'well_completion', 'paraffin', 'pumping_unit'):
        return 'medium'
    return 'low'

def ptw_status(s):
    s = str(s).strip().lower()
    if 'закрыт' in s:
        return 'closed'
    if 'отклон' in s:
        return 'rejected'
    if 'актив' in s or 'действ' in s or 'открыт' in s:
        return 'active'
    return 'pending'

ptw_path = f'{BASE}/journal_all_all_all.xlsx'
if os.path.exists(ptw_path):
    wb4 = openpyxl.load_workbook(ptw_path, read_only=True, data_only=True)
    ws = wb4['Журнал нарядов-допусков']
    ptw_rows = list(ws.iter_rows(values_only=True))[1:]
    permits = []
    for r in ptw_rows:
        seq, start, end, num, issuer, work, status = r
        if not num:
            continue
        cat = ptw_work_category(work)
        permits.append({
            'seq': int(seq) if seq else len(permits) + 1,
            'num': str(num).strip(),
            'start': None if not start or str(start).strip() == '—' else str(start).strip(),
            'end': None if not end or str(end).strip() == '—' else str(end).strip(),
            'issuer': str(issuer).strip() if issuer else '',
            'work': str(work).strip() if work and str(work) != '—' else '',
            'status': ptw_status(status),
            'category': cat,
            'risk': ptw_risk(cat),
        })
    with open(f'{OUT}/ptw.json', 'w', encoding='utf-8') as f:
        json.dump(permits, f, ensure_ascii=False, indent=0)
    print('ptw permits:', len(permits))

# ---------- Process matrix (IRD compliance by business direction) ----------
STATUS_MAP = {
    'ВН': 'VN', 'Внедрено': 'VN',
    'АКТ': 'AKT', 'Актуализировано в 2023-2025гг.': 'AKT',
    'РАЗ': 'RAZ', 'В разработке': 'RAZ',
    'НР': 'NR', 'Не разработано': 'NR',
    'НЕС': 'NES', 'Не соответствует': 'NES',
    'Н/П': 'NP', 'Не применимо': 'NP',
    '+': 'PLUS', '−': 'MINUS', '-': 'MINUS',
    'План 2026': 'PLAN26', 'План 2027': 'PLAN27',
}

def norm_matrix_status(v):
    if v is None:
        return 'EMPTY'
    s = str(v).strip()
    return STATUS_MAP.get(s, STATUS_MAP.get(s.upper(), s))

MATRIX_META = {
    'production': {
        'label': {'ru': 'Добыча', 'kz': 'Өндіру', 'en': 'Production'},
        'file': '/policies/matrices/dobycha.xlsx',
        'xlsx': 'Матрица процессов ПБ_HSE KMG_2026_Добыча.xlsx',
    },
    'refining': {
        'label': {'ru': 'Переработка', 'kz': 'Өңдеу', 'en': 'Refining'},
        'file': '/policies/matrices/pererabotka.xlsx',
        'xlsx': 'Матрица процессов ПБ_HSE KMG_2026_Переработка.xlsx',
    },
    'transport': {
        'label': {'ru': 'Транспортировка', 'kz': 'Тасымалдау', 'en': 'Transportation'},
        'file': '/policies/matrices/transport.xlsx',
        'xlsx': 'Матрица процессов ПБ_HSE KMG_2026_Транспортировка.xlsx',
    },
    'services': {
        'label': {'ru': 'Сервис', 'kz': 'Сервис', 'en': 'Services'},
        'file': '/policies/matrices/servis.xlsx',
        'xlsx': 'Матрица процессов ПБ_HSE KMG_2026_Сервис.xlsx',
    },
}

matrix_directions = []
for key, meta in MATRIX_META.items():
    path = f'{BASE}/app/{meta["xlsx"]}'
    if not os.path.exists(path):
        path = f'{BASE}/{meta["xlsx"]}'
    if not os.path.exists(path):
        continue
    ws = openpyxl.load_workbook(path, data_only=True)['Стратегия СК_ВНД']
    header_row = next(
        r for r in range(1, 8)
        if ws.cell(r, 1).value and 'бизнес-направлению' in str(ws.cell(r, 1).value).lower()
    )
    cols = []
    c = 2
    while c <= ws.max_column:
        v = ws.cell(header_row, c).value
        if v is None:
            break
        sv = str(v).strip()
        if sv.lower().startswith('внедрено') or sv.lower().startswith('актуализ'):
            break
        cols.append(sv)
        c += 1
    rows = []
    for r in range(header_row + 1, ws.max_row + 1):
        name = ws.cell(r, 1).value
        if not name:
            continue
        name = str(name).strip()
        if 'Стратегией' in name or name.startswith('Отчет'):
            continue
        statuses = [norm_matrix_status(ws.cell(r, i + 2).value) for i in range(len(cols))]
        rows.append({'name': name, 'statuses': statuses})
    matrix_directions.append({
        'id': key,
        'label': meta['label'],
        'columns': cols,
        'xlsx': meta['file'],
        'rows': rows,
    })

if matrix_directions:
    with open(f'{OUT}/policiesMatrix.json', 'w', encoding='utf-8') as f:
        json.dump({'directions': matrix_directions}, f, ensure_ascii=False, indent=2)
    print('process matrix directions:', len(matrix_directions))

# ---------- Contractor organizations (KMG-ST-3524.1, SDR register) ----------
contr_wb = openpyxl.load_workbook(f'{BASE}/Contractors information by subsidiaries_2025.xlsx', read_only=True, data_only=True)
contr_ws = contr_wb["Data of SDR's_Contractors_2025"]
contr_rows = list(contr_ws.iter_rows(values_only=True))

def contr_num(v):
    if v is None or v == '-' or v == '':
        return 0
    try:
        return int(float(v))
    except (TypeError, ValueError):
        return 0

contractor_orgs = []
cur_dir = None
cur_dzo = None
cid = 0
for row in contr_rows:
    if not row or not any(x is not None and x != '' for x in row):
        continue
    if row[0] and isinstance(row[0], str) and row[0] in ('Переработка', 'Добыча', 'Транспортировка', 'Сервис'):
        cur_dir = row[0]
        continue
    if row[0] and isinstance(row[0], (int, float)) and row[1] and isinstance(row[1], str) and row[2] is None:
        cur_dzo = str(row[1]).strip()
        continue
    if row[1] and isinstance(row[1], (int, float)) and row[2] and isinstance(row[2], str):
        cid += 1
        ns_raw = row[4]
        contractor_orgs.append({
            'id': cid,
            'dir': cur_dir,
            'dzoShort': cur_dzo,
            'name': str(row[2]).strip(),
            'activity': str(row[3]).strip() if row[3] else None,
            'ns': contr_num(ns_raw) if ns_raw not in (None, '-', '') else 0,
            'victims': contr_num(row[5]) if row[5] not in (None, '-', '') else 0,
            'fatal': contr_num(row[6]) if row[6] not in (None, '-', '') else 0,
            'workers': contr_num(row[7]),
            'wh': contr_num(row[8]),
            'km': contr_num(row[9]) if row[9] not in (None, '-', '') else 0,
            'dtp': contr_num(row[11]) if len(row) > 11 and row[11] not in (None, '-', '') else 0,
        })

contr_summary = {
    'year': 2025,
    'source': 'KMG-ST-3524.1 / Contractors information by subsidiaries_2025',
    'orgs': len(contractor_orgs),
    'workers': sum(x['workers'] for x in contractor_orgs),
    'wh': sum(x['wh'] for x in contractor_orgs),
    'ns': sum(x['ns'] for x in contractor_orgs),
    'victims': sum(x['victims'] for x in contractor_orgs),
    'fatal': sum(x['fatal'] for x in contractor_orgs),
}
with open(f'{OUT}/contractors_orgs.json', 'w', encoding='utf-8') as f:
    json.dump({'summary': contr_summary, 'orgs': contractor_orgs}, f, ensure_ascii=False, indent=0)
print('contractor orgs:', contr_summary['orgs'], 'workers:', contr_summary['workers'], 'ns:', contr_summary['ns'])

# ---------- Leading indicators (analysis appendix) ----------
leading_path = f'{BASE}/Копия Приложение Анализ по опежающим индикаторам 4 мес 2025-2026г НК КМГ к исх. письмо по списку_(9526202v1)_.xlsx'
if os.path.exists(leading_path):
    lwb = openpyxl.load_workbook(leading_path, read_only=True, data_only=True)
    lws = lwb['апрель']
    SKIP_L = {'опережающие индикаторы', 'количество'}
    KEY_MAP = [
        ('audits', 'внутренних аудитов'), ('indepAudits', 'независимыми экспертами'),
        ('findings', 'выявленных несоответствий'), ('resolved', 'устраненных несоответствий'),
        ('pab', 'ПАБ'), ('nearMiss', 'потенциально-опасных'),
        ('unsafeActs', 'опасных действий'), ('unsafeCond', 'опасных условий'),
        ('stops', 'остановов работы'), ('medical', 'медицинские пункты'),
        ('nebosh', 'NEBOSH'), ('iosh', 'IOSH'), ('defensive', 'defensive driving'),
        ('forums', 'Форумов'), ('meetings', 'совещаний'),
    ]
    WORSE_UP = {'nearMiss', 'unsafeActs', 'unsafeCond', 'stops', 'medical', 'findings'}
    blocks = []
    for col in range(1, lws.max_column + 1):
        dzo = lws.cell(9, col).value
        if not dzo or not isinstance(dzo, str):
            continue
        dzo = dzo.strip()
        if not dzo or dzo.lower() in SKIP_L:
            continue
        prev_col = curr_col = None
        for c2 in range(col, min(col + 18, lws.max_column + 1)):
            lbl = lws.cell(10, c2).value
            if lbl == '4 мес.25':
                prev_col = c2
            if lbl == '4 мес.26':
                curr_col = c2
        if prev_col and curr_col:
            blocks.append((dzo, prev_col, curr_col))
    SHORT_TO_ID = {
        'АПНЗ': 'anpz', 'АНПЗ': 'anpz', 'ПНХЗ': 'pnhz', 'ПКОП': 'pkop', 'KPI': 'kpi',
        'КазГПЗ': 'kazgpz', 'КБМ': 'kbm', 'ОМГ': 'omg', 'ЭМГ': 'emg', 'КТО': 'kto',
        'ОСК': 'osc', 'ОТК': 'otc', 'КМГИ': 'kmgeng', 'Кетеринг': 'catering',
        'ОМС': 'oms', 'ОКК': 'okk', 'ММГ': 'mmg', 'КазахойлАктобе': 'koa', 'КТМ': 'ktm',
        'Дунга': 'dunga', 'Урихтау': 'urikhtau', 'Барлау': 'barlau', 'КГМ': 'kgm',
        'УОГ': 'uog', 'ККС': 'kks', 'МЭМ': 'mem', 'КМГС': 'kmgs', 'удтв': 'udtv',
        'МТК': 'mtk', 'CASPI BITUM': 'caspi', 'Сaspi Bitum': 'caspi', 'Дрилинг': 'drilling',
        'КМТФ': 'kmtf',
    }
    by_short = {}
    by_id = {}
    for dzo, prev_col, curr_col in blocks:
        rows = []
        for row in range(11, 26):
            label = lws.cell(row, 2).value
            if not label:
                continue
            label = str(label).strip()
            if label.lower() in SKIP_L:
                continue
            kid = None
            for key, frag in KEY_MAP:
                if frag.lower() in label.lower():
                    kid = key
                    break
            if not kid:
                kid = re.sub(r'\W+', '_', label[:24]).lower()
            prev = contr_num(lws.cell(row, prev_col).value)
            curr = contr_num(lws.cell(row, curr_col).value)
            rows.append({'id': kid, 'label': label, 'prev': prev, 'curr': curr, 'worseUp': kid in WORSE_UP})
        by_short[dzo] = rows
        pid = SHORT_TO_ID.get(dzo, dzo.lower().replace(' ', '_'))
        by_id[pid] = {'short': dzo, 'rows': rows}
    with open(f'{OUT}/passportLeading.json', 'w', encoding='utf-8') as f:
        json.dump({
            'source': 'Копия Приложение Анализ по опережающим индикаторам 4 мес 2025-2026г',
            'compareMonths': 4, 'prevYear': 2025, 'currYear': 2026,
            'byShort': by_short, 'byId': by_id,
        }, f, ensure_ascii=False, indent=2)
    print('leading indicators DZO:', len(by_id))

print('DONE')
